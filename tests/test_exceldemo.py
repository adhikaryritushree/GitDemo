import openpyxl

book = openpyxl.load_workbook("C:\\Users\\RITUSHREE\\Downloads\\Demoexcel.xlsx")
sheet = book.active
#cell = sheet.cell(row=1, column=1)
#print(cell.value)
Dict = {}
maxcol = sheet.max_column
maxrow = sheet.max_row
print(maxcol, maxrow, sep="|")


def demoexcel(test):
    for i in range(1, maxrow + 1):
        if sheet.cell(row=i, column=1).value == test:
            for j in range(2, maxcol + 1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

    print(Dict)


demoexcel("testcase 1")





